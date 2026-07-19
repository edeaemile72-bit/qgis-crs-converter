# -*- coding: utf-8 -*-
"""
CRS Converter - Boîte de dialogue.

Trois onglets :
  1) Point unique
  2) Fichier CSV (import/export)
  3) Fichier Excel multi-feuilles (import/export)

La conversion s'appuie sur QgsCoordinateTransform (moteur PROJ de QGIS),
donc la précision est identique à celle des reprojections natives de QGIS.
"""

import os
import csv

from qgis.PyQt.QtCore import Qt
from qgis.PyQt.QtWidgets import (
    QDialog, QVBoxLayout, QHBoxLayout, QFormLayout, QTabWidget, QWidget,
    QLabel, QLineEdit, QPushButton, QFileDialog, QMessageBox, QComboBox,
    QTableWidget, QTableWidgetItem, QCheckBox, QProgressBar, QGroupBox,
    QHeaderView
)

from qgis.core import (
    QgsCoordinateReferenceSystem,
    QgsCoordinateTransform,
    QgsProject,
    QgsPointXY,
    QgsVectorLayer,
    QgsFeature,
    QgsGeometry,
    QgsWkbTypes,
)
from qgis.gui import QgsProjectionSelectionWidget

# Alias d'en-têtes couramment utilisés pour repérer automatiquement
# les colonnes X/Y (longitude/latitude, easting/northing, etc.)
ALIAS_X = ["x", "lon", "long", "longitude", "easting", "e"]
ALIAS_Y = ["y", "lat", "latitude", "northing", "n"]


def detecter_colonne(entetes, alias):
    for i, h in enumerate(entetes):
        if h is not None and str(h).strip().lower() in alias:
            return i
    return -1


def decimal_vers_dms(valeur, est_latitude):
    """Convertit une coordonnée décimale (degrés) en chaîne DMS.

    est_latitude=True  -> hémisphère N/S (utilisé pour Y)
    est_latitude=False -> hémisphère E/W (utilisé pour X)
    """
    hemisphere = ("N" if valeur >= 0 else "S") if est_latitude else ("E" if valeur >= 0 else "W")
    valeur_abs = abs(valeur)
    degres = int(valeur_abs)
    minutes_decimales = (valeur_abs - degres) * 60
    minutes = int(minutes_decimales)
    secondes = (minutes_decimales - minutes) * 60
    return f"{degres}°{minutes:02d}'{secondes:05.2f}\"{hemisphere}"


class CrsConverterDialog(QDialog):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Convertisseur de coordonnées (CRS)")
        self.resize(720, 560)

        layout = QVBoxLayout(self)
        self.tabs = QTabWidget()
        layout.addWidget(self.tabs)

        self._build_tab_point()
        self._build_tab_csv()
        self._build_tab_excel()
        self._build_tab_layer()

        self.tabs.addTab(self.tab_point, "Point unique")
        self.tabs.addTab(self.tab_csv, "Fichier CSV")
        self.tabs.addTab(self.tab_excel, "Fichier Excel (multi-feuilles)")
        self.tabs.addTab(self.tab_layer, "Couche vectorielle chargée")
        self.tabs.currentChanged.connect(self._on_tab_changed)

        h_aide = QHBoxLayout()
        h_aide.addStretch()
        self.btn_aide = QPushButton("Aide")
        self.btn_aide.clicked.connect(self._afficher_aide)
        h_aide.addWidget(self.btn_aide)
        layout.addLayout(h_aide)

    # ------------------------------------------------------------------
    # Utilitaires communs
    # ------------------------------------------------------------------

    def _make_crs_row(self, label_source="CRS source", label_cible="CRS cible"):
        """Crée une paire de sélecteurs de CRS (widget natif QGIS)."""
        crs_source = QgsProjectionSelectionWidget()
        crs_cible = QgsProjectionSelectionWidget()
        crs_source.setCrs(QgsCoordinateReferenceSystem("EPSG:4326"))
        crs_cible.setCrs(QgsCoordinateReferenceSystem("EPSG:32631"))
        return crs_source, crs_cible

    def _get_transform(self, crs_source_widget, crs_cible_widget):
        crs_source = crs_source_widget.crs()
        crs_cible = crs_cible_widget.crs()
        if not crs_source.isValid() or not crs_cible.isValid():
            raise ValueError("CRS source ou cible invalide.")
        return QgsCoordinateTransform(crs_source, crs_cible, QgsProject.instance())

    # ------------------------------------------------------------------
    # Onglet 1 : Point unique
    # ------------------------------------------------------------------

    def _build_tab_point(self):
        self.tab_point = QWidget()
        v = QVBoxLayout(self.tab_point)

        form = QFormLayout()
        self.point_x = QLineEdit()
        self.point_x.setPlaceholderText("ex: 2.4238  ou  392145.67")
        self.point_y = QLineEdit()
        self.point_y.setPlaceholderText("ex: 6.3703  ou  704521.12")
        form.addRow("X (longitude / easting) :", self.point_x)
        form.addRow("Y (latitude / northing) :", self.point_y)
        v.addLayout(form)

        self.point_crs_source, self.point_crs_cible = self._make_crs_row()
        grp = QGroupBox("Systèmes de coordonnées")
        gform = QFormLayout(grp)
        gform.addRow("CRS source :", self.point_crs_source)
        gform.addRow("CRS cible :", self.point_crs_cible)
        v.addWidget(grp)

        btn = QPushButton("Convertir")
        btn.clicked.connect(self._convertir_point)
        v.addWidget(btn)

        self.point_result = QLabel("")
        self.point_result.setStyleSheet("font-weight: bold; padding: 8px;")
        self.point_result.setWordWrap(True)
        v.addWidget(self.point_result)
        v.addStretch()

    def _convertir_point(self):
        try:
            x = float(self.point_x.text().replace(",", "."))
            y = float(self.point_y.text().replace(",", "."))
        except ValueError:
            QMessageBox.warning(self, "Erreur", "Veuillez saisir des coordonnées numériques valides.")
            return

        try:
            transform = self._get_transform(self.point_crs_source, self.point_crs_cible)
            point_converti = transform.transform(QgsPointXY(x, y))
        except Exception as e:
            QMessageBox.critical(self, "Erreur de conversion", str(e))
            return

        texte = (
            f"Résultat :  X = {point_converti.x():.6f}   |   Y = {point_converti.y():.6f}"
        )

        crs_cible = self.point_crs_cible.crs()
        if crs_cible.isGeographic():
            dms_x = decimal_vers_dms(point_converti.x(), est_latitude=False)
            dms_y = decimal_vers_dms(point_converti.y(), est_latitude=True)
            texte += f"\nFormat DMS :  Lon = {dms_x}   |   Lat = {dms_y}"
        else:
            texte += "\n(Le CRS cible est projeté en mètres — le format DMS ne s'applique pas.)"

        self.point_result.setText(texte)

    # ------------------------------------------------------------------
    # Onglet 2 : CSV
    # ------------------------------------------------------------------

    def _build_tab_csv(self):
        self.tab_csv = QWidget()
        v = QVBoxLayout(self.tab_csv)

        # Sélection du fichier d'entrée
        h = QHBoxLayout()
        self.csv_input_path = QLineEdit()
        self.csv_input_path.setReadOnly(True)
        btn_browse = QPushButton("Choisir le fichier CSV...")
        btn_browse.clicked.connect(self._choisir_csv_entree)
        h.addWidget(self.csv_input_path)
        h.addWidget(btn_browse)
        v.addLayout(h)

        # Colonnes X/Y détectées
        form = QFormLayout()
        self.csv_col_x = QComboBox()
        self.csv_col_y = QComboBox()
        self.csv_separateur = QComboBox()
        self.csv_separateur.addItems([", (virgule)", "; (point-virgule)", "\\t (tabulation)"])
        form.addRow("Colonne X :", self.csv_col_x)
        form.addRow("Colonne Y :", self.csv_col_y)
        form.addRow("Séparateur :", self.csv_separateur)
        v.addLayout(form)

        self.csv_crs_source, self.csv_crs_cible = self._make_crs_row()
        grp = QGroupBox("Systèmes de coordonnées")
        gform = QFormLayout(grp)
        gform.addRow("CRS source :", self.csv_crs_source)
        gform.addRow("CRS cible :", self.csv_crs_cible)
        v.addWidget(grp)

        self.csv_log_checkbox = QCheckBox("Générer un journal d'erreurs détaillé (.log)")
        v.addWidget(self.csv_log_checkbox)

        self.csv_progress = QProgressBar()
        v.addWidget(self.csv_progress)

        btn = QPushButton("Convertir et exporter le CSV")
        btn.clicked.connect(self._convertir_csv)
        v.addWidget(btn)

        self.csv_result = QLabel("")
        self.csv_result.setWordWrap(True)
        v.addWidget(self.csv_result)
        v.addStretch()

        self._csv_headers = []
        self._csv_sep_char = ","

    def _sep_char(self):
        idx = self.csv_separateur.currentIndex()
        return [",", ";", "\t"][idx]

    def _choisir_csv_entree(self):
        path, _ = QFileDialog.getOpenFileName(self, "Choisir un fichier CSV", "", "CSV (*.csv)")
        if not path:
            return
        self.csv_input_path.setText(path)

        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                sample = f.readline()
                # Détection simple du séparateur à partir de la 1re ligne
                if ";" in sample and sample.count(";") >= sample.count(","):
                    self.csv_separateur.setCurrentIndex(1)
                elif "\t" in sample:
                    self.csv_separateur.setCurrentIndex(2)
                else:
                    self.csv_separateur.setCurrentIndex(0)

                f.seek(0)
                reader = csv.reader(f, delimiter=self._sep_char())
                headers = next(reader)
        except Exception as e:
            QMessageBox.critical(self, "Erreur de lecture", str(e))
            return

        self._csv_headers = headers
        self.csv_col_x.clear()
        self.csv_col_y.clear()
        self.csv_col_x.addItems(headers)
        self.csv_col_y.addItems(headers)

        ix = detecter_colonne(headers, ALIAS_X)
        iy = detecter_colonne(headers, ALIAS_Y)
        if ix >= 0:
            self.csv_col_x.setCurrentIndex(ix)
        if iy >= 0:
            self.csv_col_y.setCurrentIndex(iy)

    def _convertir_csv(self):
        input_path = self.csv_input_path.text()
        if not input_path:
            QMessageBox.warning(self, "Erreur", "Veuillez d'abord choisir un fichier CSV.")
            return

        output_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le CSV converti", "", "CSV (*.csv)"
        )
        if not output_path:
            return

        try:
            transform = self._get_transform(self.csv_crs_source, self.csv_crs_cible)
        except Exception as e:
            QMessageBox.critical(self, "Erreur de conversion", str(e))
            return

        col_x = self.csv_col_x.currentText()
        col_y = self.csv_col_y.currentText()
        sep = self._sep_char()

        try:
            with open(input_path, newline="", encoding="utf-8-sig") as fin:
                reader = list(csv.DictReader(fin, delimiter=sep))

            total = len(reader)
            self.csv_progress.setMaximum(total if total else 1)

            fieldnames = list(reader[0].keys()) if reader else [col_x, col_y]
            crs_cible_obj = self.csv_crs_cible.crs()
            crs_cible_str = crs_cible_obj.authid()
            avec_dms = crs_cible_obj.isGeographic()

            fieldnames = fieldnames + [f"X_{crs_cible_str}", f"Y_{crs_cible_str}"]
            if avec_dms:
                fieldnames = fieldnames + ["Lon_DMS", "Lat_DMS"]

            with open(output_path, "w", newline="", encoding="utf-8") as fout:
                writer = csv.DictWriter(fout, fieldnames=fieldnames, delimiter=sep)
                writer.writeheader()
                erreurs = 0
                details_erreurs = []
                for i, row in enumerate(reader):
                    try:
                        x = float(str(row[col_x]).replace(",", "."))
                        y = float(str(row[col_y]).replace(",", "."))
                        pt = transform.transform(QgsPointXY(x, y))
                        row[f"X_{crs_cible_str}"] = pt.x()
                        row[f"Y_{crs_cible_str}"] = pt.y()
                        if avec_dms:
                            row["Lon_DMS"] = decimal_vers_dms(pt.x(), est_latitude=False)
                            row["Lat_DMS"] = decimal_vers_dms(pt.y(), est_latitude=True)
                    except Exception as e_ligne:
                        erreurs += 1
                        row[f"X_{crs_cible_str}"] = ""
                        row[f"Y_{crs_cible_str}"] = ""
                        if avec_dms:
                            row["Lon_DMS"] = ""
                            row["Lat_DMS"] = ""
                        details_erreurs.append(
                            f"Ligne {i + 2} : {col_x}={row.get(col_x)!r}, {col_y}={row.get(col_y)!r} "
                            f"— raison : {e_ligne}"
                        )
                    writer.writerow(row)
                    self.csv_progress.setValue(i + 1)

            chemin_log = None
            if self.csv_log_checkbox.isChecked():
                chemin_log = os.path.splitext(output_path)[0] + "_erreurs.log"
                with open(chemin_log, "w", encoding="utf-8") as flog:
                    flog.write(f"Journal d'erreurs — conversion CSV\n")
                    flog.write(f"Fichier source : {input_path}\n")
                    flog.write(f"Fichier converti : {output_path}\n")
                    flog.write(f"Nombre total de lignes : {total}\n")
                    flog.write(f"Nombre d'erreurs : {erreurs}\n")
                    flog.write("-" * 60 + "\n")
                    if details_erreurs:
                        flog.write("\n".join(details_erreurs))
                    else:
                        flog.write("Aucune erreur.\n")

        except Exception as e:
            QMessageBox.critical(self, "Erreur", str(e))
            return

        msg = f"Conversion terminée : {total} ligne(s) traitée(s)."
        if avec_dms:
            msg += "\nColonnes Lon_DMS / Lat_DMS ajoutées (CRS cible géographique)."
        else:
            msg += "\nPas de colonnes DMS (CRS cible projeté en mètres)."
        if erreurs:
            msg += f"\n{erreurs} ligne(s) n'ont pas pu être converties (valeurs manquantes ou invalides)."
        if chemin_log:
            msg += f"\nJournal d'erreurs enregistré : {chemin_log}"
        self.csv_result.setText(msg)
        QMessageBox.information(self, "Terminé", msg)

    # ------------------------------------------------------------------
    # Onglet 3 : Excel multi-feuilles
    # ------------------------------------------------------------------

    def _build_tab_excel(self):
        self.tab_excel = QWidget()
        v = QVBoxLayout(self.tab_excel)

        h = QHBoxLayout()
        self.xlsx_input_path = QLineEdit()
        self.xlsx_input_path.setReadOnly(True)
        btn_browse = QPushButton("Choisir le fichier Excel (.xlsx)...")
        btn_browse.clicked.connect(self._choisir_excel_entree)
        h.addWidget(self.xlsx_input_path)
        h.addWidget(btn_browse)
        v.addLayout(h)

        v.addWidget(QLabel(
            "Cochez les feuilles à convertir, vérifiez les colonnes X/Y détectées, et "
            "ajustez le CRS source si une feuille provient d'un système différent :"
        ))

        self.xlsx_table = QTableWidget(0, 5)
        self.xlsx_table.setHorizontalHeaderLabels(
            ["Inclure", "Feuille", "Colonne X", "Colonne Y", "CRS source (feuille)"]
        )
        self.xlsx_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        v.addWidget(self.xlsx_table)

        self.xlsx_crs_source, self.xlsx_crs_cible = self._make_crs_row()
        grp = QGroupBox(
            "CRS cible (unique, appliqué à toutes les feuilles) — le CRS source par "
            "défaut ci-dessous est repris pour chaque feuille au chargement, modifiable "
            "ensuite individuellement dans le tableau"
        )
        gform = QFormLayout(grp)
        gform.addRow("CRS source par défaut :", self.xlsx_crs_source)
        gform.addRow("CRS cible :", self.xlsx_crs_cible)
        v.addWidget(grp)

        self.xlsx_log_checkbox = QCheckBox("Générer un journal d'erreurs détaillé (.log)")
        v.addWidget(self.xlsx_log_checkbox)

        self.xlsx_progress = QProgressBar()
        v.addWidget(self.xlsx_progress)

        btn = QPushButton("Convertir et exporter le classeur Excel")
        btn.clicked.connect(self._convertir_excel)
        v.addWidget(btn)

        self.xlsx_result = QLabel("")
        self.xlsx_result.setWordWrap(True)
        v.addWidget(self.xlsx_result)

        self._openpyxl_ok = self._check_openpyxl()

    def _check_openpyxl(self):
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    def _choisir_excel_entree(self):
        if not self._openpyxl_ok:
            QMessageBox.critical(
                self, "Module manquant",
                "Le module 'openpyxl' n'est pas installé dans l'environnement Python de QGIS.\n\n"
                "Ouvrez le 'OSGeo4W Shell' (ou l'invite de commande QGIS) et exécutez :\n\n"
                '   python -m pip install openpyxl\n\n'
                "Puis redémarrez QGIS."
            )
            return

        path, _ = QFileDialog.getOpenFileName(self, "Choisir un fichier Excel", "", "Excel (*.xlsx)")
        if not path:
            return
        self.xlsx_input_path.setText(path)

        import openpyxl
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            QMessageBox.critical(self, "Erreur de lecture", str(e))
            return

        self.xlsx_table.setRowCount(0)
        for nom_feuille in wb.sheetnames:
            ws = wb[nom_feuille]
            try:
                first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                first_row = []
            headers = [str(h) if h is not None else "" for h in first_row]

            row_idx = self.xlsx_table.rowCount()
            self.xlsx_table.insertRow(row_idx)

            chk = QCheckBox()
            chk.setChecked(True)
            self.xlsx_table.setCellWidget(row_idx, 0, chk)
            self.xlsx_table.setItem(row_idx, 1, QTableWidgetItem(nom_feuille))

            combo_x = QComboBox()
            combo_y = QComboBox()
            combo_x.addItems(headers)
            combo_y.addItems(headers)
            ix = detecter_colonne(headers, ALIAS_X)
            iy = detecter_colonne(headers, ALIAS_Y)
            if ix >= 0:
                combo_x.setCurrentIndex(ix)
            if iy >= 0:
                combo_y.setCurrentIndex(iy)
            self.xlsx_table.setCellWidget(row_idx, 2, combo_x)
            self.xlsx_table.setCellWidget(row_idx, 3, combo_y)

            crs_feuille = QgsProjectionSelectionWidget()
            crs_feuille.setCrs(self.xlsx_crs_source.crs())
            self.xlsx_table.setCellWidget(row_idx, 4, crs_feuille)

        wb.close()

    def _convertir_excel(self):
        if not self._openpyxl_ok:
            QMessageBox.critical(self, "Module manquant", "openpyxl n'est pas installé (voir message précédent).")
            return

        input_path = self.xlsx_input_path.text()
        if not input_path:
            QMessageBox.warning(self, "Erreur", "Veuillez d'abord choisir un fichier Excel.")
            return

        output_path, _ = QFileDialog.getSaveFileName(
            self, "Enregistrer le classeur converti", "", "Excel (*.xlsx)"
        )
        if not output_path:
            return

        crs_cible_obj = self.xlsx_crs_cible.crs()
        if not crs_cible_obj.isValid():
            QMessageBox.critical(self, "Erreur de conversion", "CRS cible invalide.")
            return
        crs_cible_str = crs_cible_obj.authid()
        avec_dms = crs_cible_obj.isGeographic()

        import openpyxl
        wb = openpyxl.load_workbook(input_path, data_only=True)

        feuilles_a_traiter = []
        for row_idx in range(self.xlsx_table.rowCount()):
            chk = self.xlsx_table.cellWidget(row_idx, 0)
            if chk.isChecked():
                nom_feuille = self.xlsx_table.item(row_idx, 1).text()
                col_x = self.xlsx_table.cellWidget(row_idx, 2).currentText()
                col_y = self.xlsx_table.cellWidget(row_idx, 3).currentText()
                crs_source_feuille = self.xlsx_table.cellWidget(row_idx, 4).crs()
                feuilles_a_traiter.append((nom_feuille, col_x, col_y, crs_source_feuille))

        if not feuilles_a_traiter:
            QMessageBox.warning(self, "Erreur", "Aucune feuille cochée.")
            return

        total_lignes = 0
        total_erreurs = 0
        details_erreurs = []
        self.xlsx_progress.setMaximum(len(feuilles_a_traiter))

        for i, (nom_feuille, col_x_name, col_y_name, crs_source_feuille) in enumerate(feuilles_a_traiter):
            ws = wb[nom_feuille]
            entetes = [c.value for c in ws[1]]
            try:
                idx_x = entetes.index(col_x_name)
                idx_y = entetes.index(col_y_name)
            except ValueError:
                total_erreurs += 1
                details_erreurs.append(
                    f"Feuille '{nom_feuille}' : colonne '{col_x_name}' ou '{col_y_name}' introuvable — feuille ignorée."
                )
                continue

            if not crs_source_feuille.isValid():
                total_erreurs += 1
                details_erreurs.append(f"Feuille '{nom_feuille}' : CRS source invalide — feuille ignorée.")
                continue
            transform = QgsCoordinateTransform(crs_source_feuille, crs_cible_obj, QgsProject.instance())

            col_out_x = ws.max_column + 1
            col_out_y = ws.max_column + 2
            ws.cell(row=1, column=col_out_x, value=f"X_{crs_cible_str}")
            ws.cell(row=1, column=col_out_y, value=f"Y_{crs_cible_str}")

            if avec_dms:
                col_out_lon_dms = col_out_y + 1
                col_out_lat_dms = col_out_y + 2
                ws.cell(row=1, column=col_out_lon_dms, value="Lon_DMS")
                ws.cell(row=1, column=col_out_lat_dms, value="Lat_DMS")

            for r in range(2, ws.max_row + 1):
                x_val = ws.cell(row=r, column=idx_x + 1).value
                y_val = ws.cell(row=r, column=idx_y + 1).value
                if x_val is None or y_val is None:
                    continue
                try:
                    pt = transform.transform(QgsPointXY(float(x_val), float(y_val)))
                    ws.cell(row=r, column=col_out_x, value=pt.x())
                    ws.cell(row=r, column=col_out_y, value=pt.y())
                    if avec_dms:
                        ws.cell(row=r, column=col_out_lon_dms, value=decimal_vers_dms(pt.x(), est_latitude=False))
                        ws.cell(row=r, column=col_out_lat_dms, value=decimal_vers_dms(pt.y(), est_latitude=True))
                    total_lignes += 1
                except Exception as e_ligne:
                    total_erreurs += 1
                    details_erreurs.append(
                        f"Feuille '{nom_feuille}', ligne {r} : X={x_val!r}, Y={y_val!r} — raison : {e_ligne}"
                    )

            self.xlsx_progress.setValue(i + 1)

        try:
            wb.save(output_path)
        except Exception as e:
            QMessageBox.critical(self, "Erreur d'enregistrement", str(e))
            return

        chemin_log = None
        if self.xlsx_log_checkbox.isChecked():
            chemin_log = os.path.splitext(output_path)[0] + "_erreurs.log"
            with open(chemin_log, "w", encoding="utf-8") as flog:
                flog.write("Journal d'erreurs — conversion Excel multi-feuilles\n")
                flog.write(f"Fichier source : {input_path}\n")
                flog.write(f"Fichier converti : {output_path}\n")
                flog.write(f"Feuilles traitées : {len(feuilles_a_traiter)}\n")
                flog.write(f"Points convertis : {total_lignes}\n")
                flog.write(f"Nombre d'erreurs : {total_erreurs}\n")
                flog.write("-" * 60 + "\n")
                if details_erreurs:
                    flog.write("\n".join(details_erreurs))
                else:
                    flog.write("Aucune erreur.\n")

        msg = (
            f"Conversion terminée sur {len(feuilles_a_traiter)} feuille(s).\n"
            f"{total_lignes} point(s) converti(s), {total_erreurs} erreur(s)/ligne(s) ignorée(s).\n"
        )
        if avec_dms:
            msg += "Colonnes Lon_DMS / Lat_DMS ajoutées (CRS cible géographique).\n"
        else:
            msg += "Pas de colonnes DMS (CRS cible projeté en mètres).\n"
        if chemin_log:
            msg += f"Journal d'erreurs enregistré : {chemin_log}\n"
        msg += f"Fichier enregistré : {output_path}"
        self.xlsx_result.setText(msg)
        QMessageBox.information(self, "Terminé", msg)

    # ------------------------------------------------------------------
    # Onglet 4 : Couche vectorielle déjà chargée dans QGIS
    # ------------------------------------------------------------------

    def _build_tab_layer(self):
        self.tab_layer = QWidget()
        v = QVBoxLayout(self.tab_layer)

        v.addWidget(QLabel(
            "Convertit une couche vectorielle déjà ouverte dans QGIS (shapefile, "
            "GeoPackage, etc.) et ajoute le résultat au projet en cours, sans "
            "modifier la couche d'origine."
        ))

        h = QHBoxLayout()
        self.layer_combo = QComboBox()
        btn_refresh = QPushButton("Actualiser la liste")
        btn_refresh.clicked.connect(self._refresh_layer_list)
        h.addWidget(self.layer_combo)
        h.addWidget(btn_refresh)
        v.addLayout(h)

        self.layer_crs_source_label = QLabel("CRS source : —")
        v.addWidget(self.layer_crs_source_label)
        self.layer_combo.currentIndexChanged.connect(self._update_layer_source_label)

        grp = QGroupBox("CRS cible")
        gform = QFormLayout(grp)
        self.layer_crs_cible = QgsProjectionSelectionWidget()
        self.layer_crs_cible.setCrs(QgsCoordinateReferenceSystem("EPSG:32631"))
        gform.addRow("CRS cible :", self.layer_crs_cible)
        v.addWidget(grp)

        self.layer_progress = QProgressBar()
        v.addWidget(self.layer_progress)

        btn = QPushButton("Convertir et ajouter au projet")
        btn.clicked.connect(self._convertir_layer)
        v.addWidget(btn)

        self.layer_result = QLabel("")
        self.layer_result.setWordWrap(True)
        v.addWidget(self.layer_result)
        v.addStretch()

        self._refresh_layer_list()

    def _on_tab_changed(self, index):
        if self.tabs.widget(index) is self.tab_layer:
            self._refresh_layer_list()

    def _afficher_aide(self):
        textes_aide = {
            0: (
                "Onglet Point unique\n\n"
                "1. Saisissez X (longitude ou easting) et Y (latitude ou northing).\n"
                "2. Choisissez le CRS source (système du point saisi) et le CRS "
                "cible (système souhaité en sortie) — recherchez par nom ou code "
                "EPSG dans les sélecteurs.\n"
                "3. Cliquez sur Convertir.\n\n"
                "Si le CRS cible est géographique (ex. EPSG:4326), le résultat "
                "s'affiche aussi au format degrés-minutes-secondes (DMS)."
            ),
            1: (
                "Onglet Fichier CSV\n\n"
                "1. Choisissez votre fichier CSV : le séparateur et les colonnes "
                "X/Y sont détectés automatiquement, corrigez-les si besoin.\n"
                "2. Choisissez le CRS source et le CRS cible.\n"
                "3. Cochez 'Générer un journal d'erreurs' si vous voulez un fichier "
                ".log listant les lignes non converties.\n"
                "4. Cliquez sur Convertir et exporter le CSV.\n\n"
                "Le fichier de sortie conserve toutes vos colonnes d'origine et "
                "ajoute les colonnes converties (X/Y, et Lon_DMS/Lat_DMS si le "
                "CRS cible est géographique)."
            ),
            2: (
                "Onglet Fichier Excel (multi-feuilles)\n\n"
                "1. Choisissez votre classeur .xlsx : toutes les feuilles sont "
                "listées, avec détection automatique des colonnes X/Y.\n"
                "2. Décochez les feuilles à ignorer, corrigez les colonnes si "
                "besoin.\n"
                "3. Réglez le CRS source de chaque feuille individuellement "
                "(utile si vos feuilles viennent de systèmes différents), et le "
                "CRS cible commun en bas.\n"
                "4. Cliquez sur Convertir et exporter le classeur Excel.\n\n"
                "Nécessite le module openpyxl installé dans l'environnement "
                "Python de QGIS (voir le README du plugin)."
            ),
            3: (
                "Onglet Couche vectorielle chargée\n\n"
                "1. Sélectionnez une couche déjà ouverte dans QGIS (cliquez sur "
                "'Actualiser la liste' si elle vient d'être ajoutée).\n"
                "2. Le CRS source est détecté automatiquement depuis la couche.\n"
                "3. Choisissez le CRS cible.\n"
                "4. Cliquez sur Convertir et ajouter au projet.\n\n"
                "Une nouvelle couche mémoire reprojetée est ajoutée au projet ; "
                "la couche d'origine n'est jamais modifiée. Pour la conserver "
                "durablement, faites ensuite un clic droit sur la nouvelle "
                "couche > Exporter > Enregistrer les entités sous..."
            ),
        }
        index_actuel = self.tabs.currentIndex()
        texte = textes_aide.get(index_actuel, "Aide non disponible pour cet onglet.")
        texte += (
            "\n\n—\nCRS Converter — développé par Dr EDEA O. Emile.\n"
            "Pour l'installation, les dépendances et le détail complet des "
            "fonctionnalités, consultez le fichier README.md fourni avec le plugin."
        )
        QMessageBox.information(self, "Aide — CRS Converter", texte)

    def _refresh_layer_list(self):
        couche_actuelle = self.layer_combo.currentData()
        self.layer_combo.clear()
        for layer_id, layer in QgsProject.instance().mapLayers().items():
            if isinstance(layer, QgsVectorLayer):
                self.layer_combo.addItem(layer.name(), layer_id)
        if couche_actuelle:
            idx = self.layer_combo.findData(couche_actuelle)
            if idx >= 0:
                self.layer_combo.setCurrentIndex(idx)
        self._update_layer_source_label()

    def _update_layer_source_label(self):
        layer_id = self.layer_combo.currentData()
        layer = QgsProject.instance().mapLayer(layer_id) if layer_id else None
        if layer is not None:
            self.layer_crs_source_label.setText(f"CRS source (détecté) : {layer.crs().authid()}")
        else:
            self.layer_crs_source_label.setText("CRS source : —")

    def _convertir_layer(self):
        layer_id = self.layer_combo.currentData()
        layer = QgsProject.instance().mapLayer(layer_id) if layer_id else None
        if layer is None:
            QMessageBox.warning(
                self, "Erreur",
                "Aucune couche vectorielle sélectionnée. Cliquez sur 'Actualiser la "
                "liste' si la couche vient d'être ajoutée au projet."
            )
            return

        crs_source = layer.crs()
        crs_cible = self.layer_crs_cible.crs()
        if not crs_source.isValid() or not crs_cible.isValid():
            QMessageBox.critical(self, "Erreur de conversion", "CRS source ou cible invalide.")
            return

        transform = QgsCoordinateTransform(crs_source, crs_cible, QgsProject.instance())

        geom_type_str = QgsWkbTypes.displayString(layer.wkbType())
        nouvelle_couche = QgsVectorLayer(
            f"{geom_type_str}?crs={crs_cible.authid()}",
            f"{layer.name()}_reproj_{crs_cible.authid().replace(':', '')}",
            "memory"
        )
        pr = nouvelle_couche.dataProvider()
        pr.addAttributes(layer.fields())
        nouvelle_couche.updateFields()

        total = layer.featureCount()
        self.layer_progress.setMaximum(total if total else 1)

        nouvelles_features = []
        erreurs = 0
        for i, feat in enumerate(layer.getFeatures()):
            try:
                geom = QgsGeometry(feat.geometry())
                geom.transform(transform)
                nouveau_feat = QgsFeature(nouvelle_couche.fields())
                nouveau_feat.setAttributes(feat.attributes())
                nouveau_feat.setGeometry(geom)
                nouvelles_features.append(nouveau_feat)
            except Exception:
                erreurs += 1
            self.layer_progress.setValue(i + 1)

        pr.addFeatures(nouvelles_features)
        nouvelle_couche.updateExtents()
        QgsProject.instance().addMapLayer(nouvelle_couche)

        msg = (
            f"Couche '{nouvelle_couche.name()}' ajoutée au projet.\n"
            f"{len(nouvelles_features)} entité(s) converties de {crs_source.authid()} "
            f"vers {crs_cible.authid()}."
        )
        if erreurs:
            msg += f"\n{erreurs} entité(s) n'ont pas pu être converties."
        self.layer_result.setText(msg)
        QMessageBox.information(self, "Terminé", msg)
        self._refresh_layer_list()
